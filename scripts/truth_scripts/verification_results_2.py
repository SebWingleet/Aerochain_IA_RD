#!/usr/bin/env python3
"""
Script de validation des données LogCard extraites
Compare les résultats de l'extraction avec les données de référence (ground truth)
et génère un fichier Excel coloré avec les résultats de validation
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import os

def compare_and_create_validated_excel(extracted_json_path, ground_truth_json_path, output_dir=None):
    """
    Compare les données extraites avec les données vraies et crée un Excel coloré
    
    Args:
        extracted_json_path (str): Chemin vers le fichier JSON des données extraites
        ground_truth_json_path (str): Chemin vers le fichier JSON des données de référence
        output_dir (str): Dossier de sortie (optionnel)
    
    Returns:
        str: Chemin du fichier Excel généré
    """
    
    print("📊 VALIDATION DES DONNÉES LOGCARD")
    print("="*50)
    
    # Lecture des fichiers JSON
    print(f"📖 Lecture des données extraites: {os.path.basename(extracted_json_path)}")
    try:
        with open(extracted_json_path, 'r', encoding='utf-8') as file:
            extracted_data = json.load(file)
    except FileNotFoundError:
        print(f"❌ Fichier de données extraites non trouvé : {extracted_json_path}")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ Erreur de lecture JSON (données extraites): {e}")
        return None
    
    print(f"📖 Lecture des données de référence: {os.path.basename(ground_truth_json_path)}")
    try:
        with open(ground_truth_json_path, 'r', encoding='utf-8') as file:
            ground_truth_data = json.load(file)
    except FileNotFoundError:
        print(f"❌ Fichier ground truth non trouvé : {ground_truth_json_path}")
        print("💡 Création d'un template de fichier ground truth...")
        template_path = create_ground_truth_template(extracted_data, ground_truth_json_path)
        if template_path:
            print(f"📝 Template créé: {template_path}")
            print("⚠️  Veuillez compléter ce fichier avec les données vraies avant de relancer la validation.")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ Erreur de lecture JSON (ground truth): {e}")
        return None
    
    # Statistiques des données
    extracted_count = len(extracted_data.get('logCards', []))
    ground_truth_count = len(ground_truth_data.get('logCards', []))
    
    print(f"📊 LogCards extraites: {extracted_count}")
    print(f"📊 LogCards de référence: {ground_truth_count}")
    
    # Création du mapping des données vraies par logCard ID
    ground_truth_map = {}
    for card in ground_truth_data.get('logCards', []):
        card_id = card.get('logCard')
        if card_id is not None:
            ground_truth_map[card_id] = card.get('logCardData', {})
    
    print(f"🗂️  Mapping créé pour {len(ground_truth_map)} LogCards de référence")
    
    # Préparation des données pour l'Excel avec validation
    excel_data = []
    validation_data = []
    matched_cards = 0
    
    for card in extracted_data.get('logCards', []):
        card_id = card.get('logCard')
        extracted_card_data = card.get('logCardData', {})
        ground_truth_card_data = ground_truth_map.get(card_id, {})
        
        if ground_truth_card_data:
            matched_cards += 1
            print(f"✅ LogCard {card_id}: Données de référence trouvées")
        else:
            print(f"⚠️  LogCard {card_id}: Aucune donnée de référence")
        
        # Construction de la ligne de données
        row_data, row_validation = build_validated_row(card_id, extracted_card_data, ground_truth_card_data)
        
        excel_data.append(row_data)
        validation_data.append(row_validation)
    
    print(f"🔗 LogCards avec référence: {matched_cards}/{extracted_count}")
    
    # Créer le DataFrame
    if not excel_data:
        print("❌ Aucune donnée à traiter")
        return None
    
    df = pd.DataFrame(excel_data)
    
    # Créer le fichier Excel avec couleurs
    output_path = get_output_path(extracted_json_path, ground_truth_json_path, output_dir)
    create_colored_excel(df, validation_data, output_path, extracted_data, ground_truth_data)
    
    # Statistiques de validation
    print_validation_stats(validation_data, matched_cards, extracted_count)
    
    return output_path

def build_validated_row(card_id, extracted_data, ground_truth_data):
    """
    Construit une ligne de données et les informations de validation
    
    Args:
        card_id (int): ID de la LogCard
        extracted_data (dict): Données extraites
        ground_truth_data (dict): Données de référence
    
    Returns:
        tuple: (données_ligne, validation_ligne)
    """
    
    row_data = {
        'LogCard_ID': card_id,
        'Analyse': 'Auto-extraction',
        'Assembly_level': 'Test',
    }
    
    row_validation = {
        'LogCard_ID': 'info',  # Champ informatif
        'Analyse': 'Auto-extraction',
        'Assembly_level': 'Test',
    }
    
    # Mapping des champs principaux
    field_mappings = {
        'ATA': 'ATA',
        'Designation': 'Name', 
        'P_N': 'Manufacturer_PN',
        'S_N': 'SN',
        'Installation_Date_AC': 'install_Date_AC',
        #'TSN_AC': 'TSN_AC',
        #'CSN_AC': 'CSN_AC',
        'TSN_Part': 'TSN_Part',
        'CSN_Part': 'CSN_Part'
    }
    
    # Champs définis par le ground truth
    for field_name, gt_value in ground_truth_data.items():
        if field_name in {"TSN_AC", "CSN_AC"}:
            continue
        extracted_value = extracted_data.get(field_name)
        row_data[field_name] = normalize_value(extracted_value)
        row_validation[field_name] = validate_field(extracted_value, gt_value, field_name)
    


    
    return row_data, row_validation


def normalize_value(value):
    """
    Normalise une valeur pour la comparaison
    
    Args:
        value: Valeur à normaliser
    
    Returns:
        str: Valeur normalisée
    """
    if value is None:
        return ''
    
    # Conversion en string et nettoyage
    str_value = str(value).strip()
    
    # Cas spéciaux
    if str_value.lower() in ['null', 'none', '']:
        return ''
    
    return str_value

def validate_field(extracted_value, ground_truth_value, field_name=None):
    """
    Valide un champ et retourne le statut de validation
    
    Args:
        extracted_value: Valeur extraite
        ground_truth_value: Valeur de référence
        field_name (str): Nom du champ pour règles spécifiques
    
    Returns:
        str: Statut de validation ('correct', 'incorrect', 'no_ground_truth', 'both_empty')
    """
    extracted_norm = normalize_value(extracted_value)
    ground_truth_norm = normalize_value(ground_truth_value)
    
    if not ground_truth_norm:  # Pas de référence
        return 'no_ground_truth'
    
    if not extracted_norm and not ground_truth_norm:  # Tous deux vides
        return 'both_empty'
    
    # Règle spéciale pour le champ Name : inclusion au lieu d'égalité stricte
    if field_name == "Name":
        if ground_truth_norm.lower() in extracted_norm.lower():
            return 'correct'
        else:
            return 'incorrect'
    
    # Cas général : égalité stricte
    if extracted_norm == ground_truth_norm:
        return 'correct'
    else:
        return 'incorrect'


def validate_item_consumed(extracted_data, ground_truth_data):
    """
    Valide le champ Item Consumed qui est calculé à partir de TSN_Part et CSN_Part
    
    Args:
        extracted_data (dict): Données extraites
        ground_truth_data (dict): Données de référence
    
    Returns:
        str: Statut de validation
    """
    extracted_tsn = extracted_data.get('TSN_Part')
    extracted_csn = extracted_data.get('CSN_Part')
    ground_truth_tsn = ground_truth_data.get('TSN_Part')
    ground_truth_csn = ground_truth_data.get('CSN_Part')
    
    # Si pas de référence pour les deux champs
    if ground_truth_tsn is None and ground_truth_csn is None:
        return 'no_ground_truth'
    
    tsn_valid = validate_field(extracted_tsn, ground_truth_tsn)
    csn_valid = validate_field(extracted_csn, ground_truth_csn)
    
    # Si les deux sont corrects
    if tsn_valid == 'correct' and csn_valid == 'correct':
        return 'correct'
    # Si au moins un est incorrect
    elif tsn_valid == 'incorrect' or csn_valid == 'incorrect':
        return 'incorrect'
    # Sinon (pas de référence)
    else:
        return 'no_ground_truth'

def build_item_consumed_string(card_data):
    """
    Construit la chaîne pour Item Consumed à partir des données TSN_Part et CSN_Part
    
    Args:
        card_data (dict): Données de la LogCard
    
    Returns:
        str: Chaîne formatée
    """
    tsn_part = card_data.get('TSN_Part', '')
    csn_part = card_data.get('CSN_Part', '')
    
    if tsn_part or csn_part:
        tsn_str = str(tsn_part) if tsn_part is not None else '00:00'
        csn_str = str(csn_part) if csn_part is not None else '0'
        return f"{tsn_str} FH & {csn_str} & ?? OCY"
    else:
        return ''

def create_colored_excel(df, validation_data, output_path, extracted_data, ground_truth_data):
    """
    Crée un fichier Excel avec les couleurs de validation
    
    Args:
        df (DataFrame): Données à écrire
        validation_data (list): Informations de validation
        output_path (str): Chemin de sortie
        extracted_data (dict): Données originales extraites
        ground_truth_data (dict): Données de référence
    """
    # Définition des couleurs
    colors = {
        'correct': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),      # Vert clair
        'incorrect': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),   # Rouge clair  
        'no_ground_truth': PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"), # Orange clair
        'both_empty': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),  # Lavande
        'info': PatternFill(start_color="E0F6FF", end_color="E0F6FF", fill_type="solid")         # Bleu très clair
    }
    
    # Styles de police
    header_font = Font(bold=True, size=11)
    
    # Création du workbook
    wb = Workbook()
    
    # Feuille principale : résultats de validation
    ws_main = wb.active
    ws_main.title = "Validation_Results"
    
    # Ajout des données principales
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_main.append(r)
    
    # Formatage des headers
    for cell in ws_main[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Application des couleurs de validation
    for row_idx, validation_row in enumerate(validation_data, start=2):
        for col_idx, column_name in enumerate(df.columns, start=1):
            validation_status = validation_row.get(column_name, 'no_ground_truth')
            if validation_status in colors:
                ws_main.cell(row=row_idx, column=col_idx).fill = colors[validation_status]
    
    # Ajustement de la largeur des colonnes
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_main.column_dimensions[column_letter].width = adjusted_width
    
    # Feuille de légende
    ws_legend = wb.create_sheet("Legende")
    add_legend_sheet(ws_legend, colors, validation_data, extracted_data, ground_truth_data)
    
    # ✅ Assurer l'existence du dossier de sortie
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Sauvegarde
    wb.save(str(out_path))
    print(f"📊 Fichier Excel validé créé : {out_path}")

def add_legend_sheet(ws, colors, validation_data, extracted_data, ground_truth_data):
    """
    Ajoute une feuille de légende avec les explications
    
    Args:
        ws: Feuille de calcul
        colors (dict): Couleurs de validation
        validation_data (list): Données de validation
        extracted_data (dict): Données extraites
        ground_truth_data (dict): Données de référence
    """
    # Titre
    ws['A1'] = "LÉGENDE DES COULEURS DE VALIDATION"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Légende des couleurs
    legend_items = [
        ('A3', 'correct', 'CORRECT', 'Données extraites identiques aux données de référence'),
        ('A4', 'incorrect', 'INCORRECT', 'Données extraites différentes des données de référence'),
        ('A5', 'no_ground_truth', 'PAS DE RÉFÉRENCE', 'Aucune donnée de référence disponible pour comparaison'),
        ('A6', 'both_empty', 'TOUS DEUX VIDES', 'Champ vide dans les deux sources'),
        ('A7', 'info', 'INFORMATIF', 'Champ informatif (ID, etc.)')
    ]
    
    for cell_ref, color_key, title, description in legend_items:
        ws[cell_ref] = title
        ws[cell_ref].fill = colors[color_key]
        ws[cell_ref].font = Font(bold=True)
        
        # Description
        desc_cell = ws[cell_ref.replace('A', 'B')]
        desc_cell.value = description
    
    # Statistiques
    ws['A9'] = "STATISTIQUES"
    ws['A9'].font = Font(bold=True, size=12)
    
    extracted_count = len(extracted_data.get('logCards', []))
    ground_truth_count = len(ground_truth_data.get('logCards', []))
    
    ws['A10'] = f"LogCards extraites: {extracted_count}"
    ws['A11'] = f"LogCards de référence: {ground_truth_count}"
    
    # Calcul des statistiques de validation
    total_fields = 0
    correct_fields = 0
    incorrect_fields = 0
    no_ground_truth_fields = 0
    
    for row in validation_data:
        for field, status in row.items():
            if field != 'LogCard_ID':  # Exclure le champ ID
                total_fields += 1
                if status == 'correct':
                    correct_fields += 1
                elif status == 'incorrect':
                    incorrect_fields += 1
                elif status == 'no_ground_truth':
                    no_ground_truth_fields += 1
    
    if total_fields > 0:
        ws['A13'] = f"Total champs validés: {total_fields}"
        ws['A14'] = f"Champs corrects: {correct_fields} ({correct_fields/total_fields*100:.1f}%)"
        ws['A15'] = f"Champs incorrects: {incorrect_fields} ({incorrect_fields/total_fields*100:.1f}%)"
        ws['A16'] = f"Champs sans référence: {no_ground_truth_fields} ({no_ground_truth_fields/total_fields*100:.1f}%)"
    
    # Ajustement des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

def print_validation_stats(validation_data, matched_cards, total_cards):
    """
    Affiche les statistiques de validation dans la console
    
    Args:
        validation_data (list): Données de validation
        matched_cards (int): Nombre de LogCards avec référence
        total_cards (int): Nombre total de LogCards
    """
    total_fields = 0
    correct_fields = 0
    incorrect_fields = 0
    no_ground_truth_fields = 0
    
    for row in validation_data:
        for field, status in row.items():
            if field != 'LogCard_ID':  # Exclure le champ ID
                total_fields += 1
                if status == 'correct':
                    correct_fields += 1
                elif status == 'incorrect':
                    incorrect_fields += 1
                elif status == 'no_ground_truth':
                    no_ground_truth_fields += 1
    
    print(f"\n📊 STATISTIQUES DE VALIDATION")
    print("="*40)
    print(f"🏷️  LogCards: {matched_cards}/{total_cards} avec référence")
    
    if total_fields > 0:
        print(f"📝 Total champs: {total_fields}")
        print(f"✅ Corrects: {correct_fields} ({correct_fields/total_fields*100:.1f}%)")
        print(f"❌ Incorrects: {incorrect_fields} ({incorrect_fields/total_fields*100:.1f}%)")
        print(f"🟡 Sans référence: {no_ground_truth_fields} ({no_ground_truth_fields/total_fields*100:.1f}%)")
        
        if matched_cards > 0:
            accuracy = correct_fields / (correct_fields + incorrect_fields) * 100 if (correct_fields + incorrect_fields) > 0 else 0
            print(f"🎯 Précision: {accuracy:.1f}% (sur champs comparables)")

def create_ground_truth_template(extracted_data, output_path):
    """
    Crée un template de fichier ground truth basé sur les données extraites
    
    Args:
        extracted_data (dict): Données extraites
        output_path (str): Chemin de sortie du template
    
    Returns:
        str: Chemin du fichier créé
    """
    try:
        template = {
            "documentInfo": {
                "filename": extracted_data.get("documentInfo", {}).get("sourceMarkdown", ""),
                "totalLogCards": extracted_data.get("documentInfo", {}).get("totalLogCards", 0),
                "validationDate": "YYYY-MM-DD",
                "validatedBy": "Manual extraction - TO BE COMPLETED",
                "notes": "Template généré automatiquement - À compléter avec les vraies valeurs"
            },
            "logCards": []
        }
        
        for card in extracted_data.get('logCards', []):
            template_card = {
                "logCard": card.get('logCard'),
                "pageNumbers": card.get('pageNumbers', []),
                "logCardData": {
                    "ATA": None,  # À compléter manuellement
                    "Name": None,
                    "Manufacturer_PN": None,
                    "SN": None,
                    "install_Date_AC": None,
                    "TSN_AC": None,
                    "CSN_AC": None,
                    "TSN_Part": None,
                    "CSN_Part": None,
                    "Inventory_lifed_components": None
                },
                "validation_status": "to_be_verified"
            }
            template['logCards'].append(template_card)
        
        with open(output_path, 'w', encoding='utf-8') as file:
            json.dump(template, file, indent=2, ensure_ascii=False)
        
        return output_path
        
    except Exception as e:
        print(f"❌ Erreur lors de la création du template: {e}")
        return None

def get_output_path(extracted_json_path, ground_truth_json_path, output_dir=None):
    """
    Génère le chemin de sortie pour le fichier Excel
    
    Args:
        extracted_json_path (str): Chemin du fichier extrait
        ground_truth_json_path (str): Chemin du fichier de référence
        output_dir (str): Dossier de sortie personnalisé
    
    Returns:
        str: Chemin de sortie
    """
    if output_dir:
        output_directory = Path(output_dir)
    else:
        # Utiliser le dossier parent du fichier extrait
        extracted_path_obj = Path(extracted_json_path)
        output_directory = extracted_path_obj.parent.parent / "validation_results"
    
    output_directory.mkdir(exist_ok=True)
    
    # Nom du fichier basé sur les fichiers d'entrée
    extracted_name = Path(extracted_json_path).stem
    output_filename = "validation.xlsx"
    
    return output_directory / output_filename

def main():
    """Interface CLI pour le script de validation"""
    
    parser = argparse.ArgumentParser(description="Validation des données LogCard extraites")
    parser.add_argument('--extracted', required=True, help="Chemin vers le fichier JSON des données extraites")
    parser.add_argument('--ground-truth', required=True, help="Chemin vers le fichier JSON des données de référence")
    parser.add_argument('--output-dir', help="Dossier de sortie (optionnel)")
    
    args = parser.parse_args()
    
    # Vérifier les fichiers d'entrée
    if not os.path.exists(args.extracted):
        print(f"❌ Fichier de données extraites non trouvé: {args.extracted}")
        return
    
    if not os.path.exists(args.ground_truth):
        print(f"❌ Fichier de données de référence non trouvé: {args.ground_truth}")
        return
    
    try:
        output_file = compare_and_create_validated_excel(
            extracted_json_path=args.extracted,
            ground_truth_json_path=args.ground_truth,
            output_dir=args.output_dir
        )
        
        if output_file:
            print(f"\n🎉 Validation terminée avec succès!")
            print(f"📁 Fichier Excel: {output_file}")
        else:
            print("\n❌ Validation échouée")
            
    except Exception as e:
        print(f"\n💥 Erreur inattendue: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()


# python verification_results_2.py --extracted "C:\Users\lilia\Desktop\freelance\clients\Wingleet\Projet_DEV_Lilian\scripts\main_scripts\WORKFLOW_RESULTS\workflow_LOGCARDS-INVENTORYLOGBOOKDataSet_ocr_result_20250825_115159\phase2_logcard\LOGCARDS-INVENTORYLOGBOOKDataSet_ocr_result_logcards.json" --ground-truth "C:\Users\lilia\Desktop\freelance\clients\Wingleet\Projet_DEV_Lilian\INPUT_DOCS\LOG_CARDS_INVENTORY_LOG_BOOK_ground_truth.json"