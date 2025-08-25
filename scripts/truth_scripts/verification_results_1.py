import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def compare_and_create_validated_excel():
    """
    Compare les données extraites avec les données vraies et crée un Excel coloré
    """
    # Chemins des fichiers
    extracted_json_path = "/Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/TEST/workflow_LOGCARDS-INVENTORYLOGBOOK6pages_20250731_171208/phase2_logcard/LOGCARDS-INVENTORYLOGBOOK6pages_ocr_result_logcards.json"
    ground_truth_json_path = "/Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/INPUT_DOCS/LOG_CARDS_INVENTORY_LOG_BOOK_ground_truth.json"
    
    # Lecture des fichiers JSON
    with open(extracted_json_path, 'r', encoding='utf-8') as file:
        extracted_data = json.load(file)
    
    try:
        with open(ground_truth_json_path, 'r', encoding='utf-8') as file:
            ground_truth_data = json.load(file)
    except FileNotFoundError:
        print(f"Fichier ground truth non trouvé : {ground_truth_json_path}")
        print("Création d'un template de fichier ground truth...")
        create_ground_truth_template(extracted_data, ground_truth_json_path)
        return None
    
    # Création du mapping des données vraies par logCard ID
    ground_truth_map = {}
    for card in ground_truth_data.get('logCards', []):
        card_id = card.get('logCard')
        ground_truth_map[card_id] = card.get('logCardData', {})
    
    # Préparation des données pour l'Excel avec validation
    excel_data = []
    validation_data = []  # Pour stocker les informations de validation
    
    for card in extracted_data.get('logCards', []):
        card_id = card.get('logCard')
        extracted_card_data = card.get('logCardData', {})
        ground_truth_card_data = ground_truth_map.get(card_id, {})
        
        # Construction de la ligne de données
        row_data, row_validation = build_validated_row(extracted_card_data, ground_truth_card_data)
        
        excel_data.append(row_data)
        validation_data.append(row_validation)
    
    # Création du DataFrame
    df = pd.DataFrame(excel_data)
    
    # Création du fichier Excel avec couleurs
    output_path = get_output_path(extracted_json_path, "_validated.xlsx")
    create_colored_excel(df, validation_data, output_path)
    
    # Statistiques de validation
    print_validation_stats(validation_data)
    
    return output_path

def build_validated_row(extracted_data, ground_truth_data):
    """
    Construit une ligne de données et les informations de validation
    """
    fields_mapping = {
        'ATA': lambda x: f"{x}-??" if x else '',
        'Kardex No': lambda x: f"{extracted_data.get('ATA', '')}-??-????" if extracted_data.get('ATA') else '',
        'Kardex designation/ Function': lambda x: f"{x} ??" if x else '',
        'Designation': lambda x: x or '',
        'P/N': lambda x: x or '',
        'S/N': lambda x: x or '',
        'Installation Date A/C': lambda x: x or '',
        'TSN A/C': lambda x: x or '',
        'CSN A/C': lambda x: x or '',
        'Item Consumed (at installation).NEW': lambda _: build_item_consumed_string(extracted_data),
        'MONITORING': lambda _: 'HT LLP' if extracted_data.get('Inventory_lifed_components') else 'O/C'
    }
    
    row_data = {
        'Analyse': '',
        'Assembly level': '',
    }
    
    row_validation = {
        'Analyse': 'no_ground_truth',  # Pas de données de référence
        'Assembly level': 'no_ground_truth',
    }
    
    # Mapping des champs principaux
    field_mappings = {
        'ATA': 'ATA',
        'Designation': 'Name',
        'P/N': 'Manufacturer_PN',
        'S/N': 'SN',
        'Installation Date A/C': 'install_Date_AC',
        'TSN A/C': 'TSN_AC',
        'CSN A/C': 'CSN_AC'
    }
    
    for excel_field, json_field in field_mappings.items():
        extracted_value = extracted_data.get(json_field)
        ground_truth_value = ground_truth_data.get(json_field)
        
        # Traitement spécial pour ATA (ajouter "-??")
        if excel_field == 'ATA' and extracted_value:
            extracted_value = f"{extracted_value}-??"
        if excel_field == 'ATA' and ground_truth_value:
            ground_truth_value = f"{ground_truth_value}-??"
        
        row_data[excel_field] = str(extracted_value) if extracted_value is not None else ''
        row_validation[excel_field] = validate_field(extracted_value, ground_truth_value)
    
    # Champs calculés
    row_data['Kardex No'] = f"{extracted_data.get('ATA', '')}-??-????" if extracted_data.get('ATA') else ''
    row_data['Kardex designation/ Function'] = f"{extracted_data.get('Name', '')} ??" if extracted_data.get('Name') else ''
    row_data['F.I.N. Code'] = ''
    row_data['Zone'] = ''
    row_data['ACCESS'] = ''
    row_data['Item Consumed (at installation).NEW'] = build_item_consumed_string(extracted_data)
    row_data['Item Consumed (at installation).Overhaul'] = ''
    row_data['Item Consumed (at installation).Maintenance'] = ''
    row_data['Item Consumed (at installation).Inspection'] = ''
    row_data['MONITORING'] = 'HT LLP' if extracted_data.get('Inventory_lifed_components') else 'O/C'
    
    # Validation des champs calculés (basée sur les champs sources)
    row_validation['Kardex No'] = validate_field(extracted_data.get('ATA'), ground_truth_data.get('ATA'))
    row_validation['Kardex designation/ Function'] = validate_field(extracted_data.get('Name'), ground_truth_data.get('Name'))
    row_validation['F.I.N. Code'] = 'no_ground_truth'
    row_validation['Zone'] = 'no_ground_truth'
    row_validation['ACCESS'] = 'no_ground_truth'
    row_validation['Item Consumed (at installation).NEW'] = validate_item_consumed(extracted_data, ground_truth_data)
    row_validation['Item Consumed (at installation).Overhaul'] = 'no_ground_truth'
    row_validation['Item Consumed (at installation).Maintenance'] = 'no_ground_truth'
    row_validation['Item Consumed (at installation).Inspection'] = 'no_ground_truth'
    row_validation['MONITORING'] = validate_field(extracted_data.get('Inventory_lifed_components'), ground_truth_data.get('Inventory_lifed_components'))
    
    return row_data, row_validation

def validate_field(extracted_value, ground_truth_value):
    """
    Valide un champ et retourne le statut de validation
    """
    if ground_truth_value is None:
        return 'no_ground_truth'
    
    # Conversion en string pour comparaison
    extracted_str = str(extracted_value) if extracted_value is not None else ''
    ground_truth_str = str(ground_truth_value) if ground_truth_value is not None else ''
    
    if extracted_str == ground_truth_str:
        return 'correct'
    else:
        return 'incorrect'

def validate_item_consumed(extracted_data, ground_truth_data):
    """
    Valide le champ Item Consumed qui est calculé
    """
    extracted_tsn = extracted_data.get('TSN_Part')
    extracted_csn = extracted_data.get('CSN_Part')
    ground_truth_tsn = ground_truth_data.get('TSN_Part')
    ground_truth_csn = ground_truth_data.get('CSN_Part')
    
    if ground_truth_tsn is None and ground_truth_csn is None:
        return 'no_ground_truth'
    
    tsn_valid = validate_field(extracted_tsn, ground_truth_tsn)
    csn_valid = validate_field(extracted_csn, ground_truth_csn)
    
    if tsn_valid == 'correct' and csn_valid == 'correct':
        return 'correct'
    elif tsn_valid == 'no_ground_truth' or csn_valid == 'no_ground_truth':
        return 'no_ground_truth'
    else:
        return 'incorrect'

def build_item_consumed_string(card_data):
    """
    Construit la chaîne pour Item Consumed
    """
    tsn_part = card_data.get('TSN_Part', '')
    csn_part = card_data.get('CSN_Part', '')
    
    if tsn_part or csn_part:
        tsn_str = str(tsn_part) if tsn_part is not None else ''
        csn_str = str(csn_part) if csn_part is not None else ''
        return f"{tsn_str} FH & {csn_str} & ?? OCY"
    else:
        return ''

def create_colored_excel(df, validation_data, output_path):
    """
    Crée un fichier Excel avec les couleurs de validation
    """
    # Définition des couleurs
    colors = {
        'correct': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),      # Vert clair
        'incorrect': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),   # Rouge clair
        'no_ground_truth': PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid") # Orange clair
    }
    
    # Création du workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"
    
    # Ajout des données
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Application des couleurs
    for row_idx, validation_row in enumerate(validation_data, start=2):  # Start=2 car ligne 1 = headers
        for col_idx, column_name in enumerate(df.columns, start=1):
            validation_status = validation_row.get(column_name, 'no_ground_truth')
            if validation_status in colors:
                ws.cell(row=row_idx, column=col_idx).fill = colors[validation_status]
    
    # Ajustement de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde
    wb.save(output_path)
    print(f"Fichier Excel validé créé : {output_path}")

def print_validation_stats(validation_data):
    """
    Affiche les statistiques de validation
    """
    total_fields = 0
    correct_fields = 0
    incorrect_fields = 0
    no_ground_truth_fields = 0
    
    for row in validation_data:
        for field, status in row.items():
            total_fields += 1
            if status == 'correct':
                correct_fields += 1
            elif status == 'incorrect':
                incorrect_fields += 1
            elif status == 'no_ground_truth':
                no_ground_truth_fields += 1
    
    print(f"\n=== STATISTIQUES DE VALIDATION ===")
    print(f"Total de champs: {total_fields}")
    print(f"Champs corrects (vert): {correct_fields} ({correct_fields/total_fields*100:.1f}%)")
    print(f"Champs incorrects (rouge): {incorrect_fields} ({incorrect_fields/total_fields*100:.1f}%)")
    print(f"Champs sans référence (orange): {no_ground_truth_fields} ({no_ground_truth_fields/total_fields*100:.1f}%)")

def create_ground_truth_template(extracted_data, output_path):
    """
    Crée un template de fichier ground truth basé sur les données extraites
    """
    template = {
        "documentInfo": {
            "filename": extracted_data.get("documentInfo", {}).get("filename", ""),
            "totalLogCards": extracted_data.get("documentInfo", {}).get("totalLogCards", 0),
            "validationDate": "2025-07-31",
            "validatedBy": "Manual extraction - TO BE COMPLETED"
        },
        "logCards": []
    }
    
    for card in extracted_data.get('logCards', []):
        template_card = {
            "logCard": card.get('logCard'),
            "pageNumbers": card.get('pageNumbers', []),
            "logCardData": {
                "ATA": None,  # À compléter manuellement
                "Name": None,
                "Manufacturer_PN": None,
                "SN": None,
                "install_Date_AC": None,
                "TSN_AC": None,
                "CSN_AC": None,
                "TSN_Part": None,
                "CSN_Part": None,
                "Inventory_lifed_components": None
            },
            "validation_status": "to_be_verified"
        }
        template.logCards.append(template_card)
    
    with open(output_path, 'w', encoding='utf-8') as file:
        json.dump(template, file, indent=2, ensure_ascii=False)
    
    print(f"Template de ground truth créé : {output_path}")
    print("Veuillez compléter ce fichier avec les données vraies avant de relancer la validation.")

def get_output_path(input_path, suffix):
    """
    Génère le chemin de sortie
    """
    input_path_obj = Path(input_path)
    parent_dir = input_path_obj.parent.parent
    output_dir = parent_dir / "results"
    output_dir.mkdir(exist_ok=True)
    
    filename = input_path_obj.stem.replace('_logcards', suffix)
    return output_dir / filename

if __name__ == "__main__":
    try:
        output_file = compare_and_create_validated_excel()
        if output_file:
            print("Validation terminée avec succès!")
    except Exception as e:
        print(f"Erreur : {e}")
        import traceback
        traceback.print_exc()