# /Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/scripts/xml_extraction/xml_extract_5.py

import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
import os

def xml_to_excel(xml_file_path, excel_file_path=None, append_to_existing=False):
    """
    Convertit un fichier XML avec structure RalWebDataTable en fichier Excel
    
    Args:
        xml_file_path (str): Chemin vers le fichier XML source
        excel_file_path (str): Chemin de sortie pour le fichier Excel (optionnel)
        append_to_existing (bool): Si True, ajoute aux donn√©es existantes
    
    Returns:
        str: Chemin du fichier Excel cr√©√©/modifi√©
    """
    
    # Si aucun chemin de sortie n'est sp√©cifi√©, cr√©er un nom bas√© sur le fichier XML
    if excel_file_path is None:
        base_name = os.path.splitext(xml_file_path)[0]
        excel_file_path = f"{base_name}_converted.xlsx"
    
    try:
        # Parser le fichier XML
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        
        # Liste pour stocker toutes les donn√©es
        data_list = []
        
        # Parcourir tous les √©l√©ments RalWebDataTable
        for table in root.findall('RalWebDataTable'):
            # Dictionnaire pour une ligne de donn√©es
            row_data = {}
            
            # Extraire toutes les donn√©es de chaque √©l√©ment
            for child in table:
                # Nettoyer le nom de la colonne et la valeur
                column_name = child.tag.strip()
                value = child.text.strip() if child.text else ""
                row_data[column_name] = value
            
            data_list.append(row_data)
        
        # Cr√©er un DataFrame pandas
        df = pd.DataFrame(data_list)
        
        # Nettoyer et formater les donn√©es
        df = clean_and_format_dataframe(df)
        
        # Transformer vers le format personnalis√©
        transformed_df = transform_to_custom_format(df)
        
        # Gestion du fichier Excel selon le mode
        if append_to_existing and os.path.exists(excel_file_path):
            # Mode ajout : ajouter aux donn√©es existantes
            result = append_to_existing_excel(transformed_df, excel_file_path)
            if not result:
                return None
        else:
            # Mode cr√©ation : cr√©er un nouveau fichier
            create_new_excel_file(transformed_df, df, excel_file_path)
        
        print(f"‚úÖ Conversion r√©ussie!")
        print(f"üìÅ Fichier Excel: {excel_file_path}")
        if append_to_existing:
            print(f"üìä Nouvelles lignes ajout√©es: {len(transformed_df)}")
        else:
            print(f"üìä Feuille 'Aircraft_Data_Custom': {len(transformed_df)} lignes, {len(transformed_df.columns)} colonnes")
            print(f"üìä Feuille 'Aircraft_Data_Original': {len(df)} lignes, {len(df.columns)} colonnes")
        
        return excel_file_path
        
    except ET.ParseError as e:
        print(f"‚ùå Erreur lors du parsing XML: {e}")
        return None
    except Exception as e:
        print(f"‚ùå Erreur g√©n√©rale: {e}")
        return None

def clean_and_format_dataframe(df):
    """
    Nettoie et formate le DataFrame
    
    Args:
        df (pandas.DataFrame): DataFrame √† nettoyer
    
    Returns:
        pandas.DataFrame: DataFrame nettoy√©
    """
    
    # Colonnes de dates √† traiter
    date_columns = ['ManufactureDate', 'InstallationDate', 'FirstInstallationDate', 'ExpiryDate']
    
    for col in date_columns:
        if col in df.columns:
            # Convertir les dates au format datetime
            df[col] = pd.to_datetime(df[col], format='%d.%m.%Y', errors='coerce')
    
    # Colonnes num√©riques √† traiter
    numeric_columns = [
        'AircraftSerialNumber', 'ModelTreeLevel', 'ATAChapter',
        'HigherAssemblyAgeingAtFitInCycles', 'ComponentAgeingatInstallationinCycles',
        'HigherAssemblyCurrentCycles', 'ComponentCurrentCycles'
    ]
    
    for col in numeric_columns:
        if col in df.columns:
            # Convertir en num√©rique, remplacer les erreurs par NaN
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Remplacer les cha√Ænes vides par NaN pour une meilleure lisibilit√©
    df = df.replace('', pd.NA)
    
    return df

def transform_to_custom_format(df):
    """
    Transforme le DataFrame selon le format demand√© avec les nouvelles colonnes
    
    Args:
        df (pandas.DataFrame): DataFrame source avec les donn√©es XML
    
    Returns:
        pandas.DataFrame: DataFrame transform√© avec les nouvelles colonnes
    """
    
    # Cr√©er un nouveau DataFrame avec les colonnes demand√©es
    transformed_data = []
    
    for _, row in df.iterrows():
        # Extraire Location et ATAChapter
        location = str(row.get('Location', '')) if pd.notna(row.get('Location')) else ''
        ata_chapter_xml = row.get('ATAChapter', '')
        model_tree_level_raw = row.get('ModelTreeLevel', 0)

        # Nettoyer ModelTreeLevel pour √©viter les erreurs NaT/NaN
        model_tree_level = 0
        if pd.notna(model_tree_level_raw):
            try:
                model_tree_level = int(float(model_tree_level_raw))
            except (ValueError, TypeError):
                model_tree_level = 0

        # Calculer Assembly Level selon les r√®gles
        assembly_level = ''
        if model_tree_level > 0 : 
            print(f"ModelTreeLevel : {model_tree_level}")
            if model_tree_level == 1:
                assembly_level = 1
            elif model_tree_level == 2:
                assembly_level = 11
            elif model_tree_level >= 3:
                assembly_level = 111
        
        # Calculer Kardex No bas√© sur Location + "00"
        kardex_no = ''
        if location.endswith('-00'):
            # Si Location se termine par "-00", garder tel quel
            kardex_no = location
        else: 
            # Sinon, ajouter "00" √† la fin
            kardex_no = f"{location}00"

        
        
        # Calculer ATA : utiliser ATAChapter si disponible, sinon extraire de Location
        ata_value = ''
        if pd.notna(ata_chapter_xml) and str(ata_chapter_xml).strip() != '':
            # ATAChapter est disponible
            ata_chapter = str(int(float(ata_chapter_xml))) if ata_chapter_xml != '' else ''
            ata_value = f"{ata_chapter}-00" if ata_chapter else ''
        elif location:
            # ATAChapter est vide, extraire de Location
            # Exemple: "32-30-02" ‚Üí "32"
            ata_from_location = location.split('-')[0] if '-' in location else location
            ata_value = ata_from_location
        
        # Extraire TSN et CSN Part (donn√©es √† l'installation)
        tsn_part = row.get('ComponentAgeingatInstallationinHours', '')
        csn_part = row.get('ComponentAgeingatInstallationinCycles', '')
        
        # Calculer "Item Consumed (at installation) NEW"
        item_consumed_new = ''
        if pd.notna(tsn_part) and str(tsn_part).strip() != '' and str(tsn_part) != '0:00':
            if pd.notna(csn_part) and str(csn_part).strip() != '' and str(csn_part) != '0.00':
                item_consumed_new = f"{tsn_part} FH & {csn_part} FC"
            else:
                item_consumed_new = f"{tsn_part} FH"
        elif pd.notna(csn_part) and str(csn_part).strip() != '' and str(csn_part) != '0.00':
            item_consumed_new = f"{csn_part} FC"
        
        # Extraire les autres donn√©es importantes
        installation_date = row.get('FirstInstallationDate', '')
        tsn_ac = row.get('HigherAssemblyAgeingAtFitInHours', '')
        csn_ac = row.get('HigherAssemblyAgeingAtFitInCycles', '')
        
        # Construire la nouvelle ligne avec toutes les colonnes dans l'ordre
        new_row = {
            'Analyse': '',  # Colonne A
            'Assembly level': assembly_level,  # Colonne B
            'ATA': ata_value,  # Colonne C
            'Kardex No': kardex_no,  # Colonne D
            'Kardex designation / Function' : row.get('InstalledPartDescription', ''),  # Colonne E
            'Designation': row.get('InstalledPartDescription', ''),  # Colonne F
            'F.I.N Code' : '', # Trouver dans ....... Colonne G
            'Zone' : row.get('Position', ''), # Colonne H
            'Access' : '', # Trouver dans ....... Colonne I
            'P_N': row.get('InstalledManufacturerPartNumber', ''),  # Colonne J
            'S_N': row.get('InstalledSerialNumber', ''),  # Colonne K
            'Installation_Date_AC': installation_date,  # Colonne L
            'TSN_AC': tsn_ac,  # Colonne M
            'CSN_AC': csn_ac,  # Colonne N
            'Item Consumed (at installation) NEW': item_consumed_new,  # Colonne O
            'Item Consumed (at installation) Overhaul': '',  # Colonne P
            'Item Consumed (at installation) Maintenance': '',  # Colonne Q
            'Item Consumed (at installation) Inspection': '',  # Colonne R
            'Monitoring': '',  # Trouver dans document Maintenance Status Colonne S
            'JAA/EASA Certificate': '',  # Colonne T
            'Item Monitoring : New (at installation) Task N': '',  # Colonne U
            'Item Monitoring : New (at installation) Authorized': ''  # Colonne V
        }
        
        transformed_data.append(new_row)
    
    # Cr√©er le nouveau DataFrame
    transformed_df = pd.DataFrame(transformed_data)
    
    # Nettoyer les valeurs vides mais garder les cha√Ænes vides pour les colonnes vides volontaires
    for col in transformed_df.columns:
        if col in ['Analyse', 'Assembly level', 'Item Consumed (at installation) Overhaul', 
                   'Item Consumed (at installation) Maintenance', 'Item Consumed (at installation) Inspection',
                   'Monitoring', 'JAA/EASA Certificate', 'Item Monitoring : New (at installation) Task N',
                   'Item Monitoring : New (at installation) Authorized']:
            # Garder ces colonnes vides
            continue
        else:
            # Pour les autres colonnes, remplacer les cha√Ænes vides par NaN seulement si n√©cessaire
            transformed_df[col] = transformed_df[col].replace('', pd.NA)
    
    return transformed_df

def append_to_existing_excel(new_data_df, excel_file_path):
    """
    Ajoute de nouvelles donn√©es √† un fichier Excel existant
    
    Args:
        new_data_df (pandas.DataFrame): Nouvelles donn√©es √† ajouter
        excel_file_path (str): Chemin vers le fichier Excel existant
    
    Returns:
        bool: True si succ√®s, False sinon
    """
    
    try:
        from openpyxl import load_workbook
        
        print(f"üìñ Lecture du fichier existant: {excel_file_path}")
        
        # Charger le workbook existant
        workbook = load_workbook(excel_file_path, keep_vba=True)  # keep_vba=True pour les fichiers .xlsm
        
        # Chercher la feuille "KARDEX"
        target_sheet_name = "KARDEX"
        if target_sheet_name in workbook.sheetnames:
            worksheet = workbook[target_sheet_name]
            print(f"üìÑ Feuille trouv√©e: {target_sheet_name}")
        else:
            print(f"‚ö†Ô∏è  Feuille '{target_sheet_name}' non trouv√©e!")
            print(f"üìã Feuilles disponibles: {workbook.sheetnames}")
            # Utiliser la premi√®re feuille comme fallback
            sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            print(f"üìÑ Utilisation de la feuille par d√©faut: {sheet_name}")
        
        # Trouver la derni√®re ligne avec des donn√©es
        last_row = worksheet.max_row
        
        # V√©rifier s'il y a des donn√©es existantes (en ignorant l'en-t√™te)
        while last_row > 1:
            # V√©rifier si la ligne contient des donn√©es (au moins une cellule non vide)
            has_data = False
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=last_row, column=col).value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_data = True
                    break
            
            if has_data:
                break
            last_row -= 1
        
        print(f"üìç Derni√®re ligne avec donn√©es: {last_row}")
        
        # Position pour commencer √† ajouter les nouvelles donn√©es
        start_row = last_row + 1
        
        print(f"üìù Ajout des nouvelles donn√©es √† partir de la ligne: {start_row}")
        
        # Debug: Afficher les premi√®res lignes de donn√©es √† ajouter
        print(f"\nüîç Aper√ßu des donn√©es √† ajouter:")
        for i, (_, row) in enumerate(new_data_df.head(3).iterrows()):
            print(f"  Ligne {i+1}:")
            print(f"    ATA: {row.get('ATA', 'N/A')}")
            print(f"    Kardex No: {row.get('Kardex No', 'N/A')}")
            print(f"    Designation: {row.get('Designation', 'N/A')}")
            print(f"    P_N: {row.get('P_N', 'N/A')}")
            print(f"    S_N: {row.get('S_N', 'N/A')}")
            print(f"    Installation_Date_AC: {row.get('Installation_Date_AC', 'N/A')}")
        
        print(f"\nüìä Nombre total de colonnes √† ajouter: {len(new_data_df.columns)}")
        print(f"üìã Colonnes: {list(new_data_df.columns)}")
        
        # Ajouter les nouvelles donn√©es ligne par ligne
        for idx, (_, row) in enumerate(new_data_df.iterrows(), start=start_row):
            for col_idx, value in enumerate(row.values, start=1):
                # Convertir les valeurs pandas NaN en None pour Excel
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    # Convertir les dates pandas en dates Python
                    value = value.to_pydatetime().date()
                
                worksheet.cell(row=idx, column=col_idx, value=value)
        
        # Sauvegarder le fichier
        workbook.save(excel_file_path)
        workbook.close()
        
        print(f"‚úÖ {len(new_data_df)} nouvelles lignes ajout√©es avec succ√®s!")
        
        return True
        
    except Exception as e:
        print(f"‚ùå Erreur lors de l'ajout aux donn√©es existantes: {e}")
        return False

def create_new_excel_file(transformed_df, original_df, excel_file_path):
    """
    Cr√©e un nouveau fichier Excel avec les donn√©es transform√©es
    
    Args:
        transformed_df (pandas.DataFrame): Donn√©es transform√©es
        original_df (pandas.DataFrame): Donn√©es originales
        excel_file_path (str): Chemin de sortie
    """
    
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        # Feuille avec les donn√©es transform√©es (format demand√©)
        transformed_df.to_excel(writer, sheet_name='Aircraft_Data_Custom', index=False)
        
        # Feuille avec les donn√©es originales (pour r√©f√©rence)
        original_df.to_excel(writer, sheet_name='Aircraft_Data_Original', index=False)
        
        # Obtenir les feuilles de calcul pour le formatage
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # D√©finir une largeur maximale raisonnable
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

def xml_to_existing_kardex(xml_file_path, kardex_file_path):
    """
    Fonction sp√©cialis√©e pour ajouter des donn√©es XML au fichier Kardex existant
    
    Args:
        xml_file_path (str): Chemin vers le fichier XML
        kardex_file_path (str): Chemin vers le fichier Kardex (.xlsm)
    
    Returns:
        bool: True si succ√®s, False sinon
    """
    
    print(f"üéØ Mode Kardex: Ajout au fichier existant")
    print(f"üìÅ XML source: {xml_file_path}")
    print(f"üìÅ Kardex cible: {kardex_file_path}")
    
    # V√©rifier que les fichiers existent
    if not os.path.exists(xml_file_path):
        print(f"‚ùå Fichier XML non trouv√©: {xml_file_path}")
        return False
    
    if not os.path.exists(kardex_file_path):
        print(f"‚ùå Fichier Kardex non trouv√©: {kardex_file_path}")
        return False
    
    # Utiliser la fonction principale avec l'option append
    result = xml_to_excel(xml_file_path, kardex_file_path, append_to_existing=True)
    
    return result is not None

def analyze_xml_structure(xml_file_path):
    """
    Analyse la structure du fichier XML et affiche des informations utiles
    
    Args:
        xml_file_path (str): Chemin vers le fichier XML
    """
    
    try:
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        
        print("üîç Analyse de la structure XML:")
        print(f"√âl√©ment racine: {root.tag}")
        
        # Compter les tables
        tables = root.findall('RalWebDataTable')
        print(f"Nombre de RalWebDataTable: {len(tables)}")
        
        # Analyser la premi√®re table pour voir la structure
        if tables:
            first_table = tables[0]
            print(f"\nüìã Colonnes disponibles ({len(first_table)} au total):")
            
            for i, child in enumerate(first_table, 1):
                value = child.text if child.text else "(vide)"
                print(f"  {i:2d}. {child.tag:35} = {value}")
        
    except Exception as e:
        print(f"‚ùå Erreur lors de l'analyse: {e}")

def main():
    """
    Fonction principale pour ex√©cution en tant que script
    """
    
    # Configuration des chemins
    xml_file = "/Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/INPUT_DOCS/xml/H160-MIS-DATA-PACK-1054-Applied Configuration.xml"  # Remplacez par le chemin de votre fichier XML
    kardex_file = "/Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/Output_docs/10_KARDEX_H160_MSN1054_V3_6.xlsm"
    
    print("=" * 80)
    print("üöÅ CONVERTISSEUR XML VERS KARDEX H160")
    print("=" * 80)
    
    # V√©rifier si le fichier XML existe
    if not os.path.exists(xml_file):
        print(f"‚ùå Fichier XML non trouv√©: {xml_file}")
        print("üí° Modifiez la variable 'xml_file' avec le bon chemin")
        return
    
    # Analyser la structure (optionnel)
    print("üîç ANALYSE DE LA STRUCTURE XML")
    print("-" * 40)
    analyze_xml_structure(xml_file)
    
    print("\nüîÑ CONVERSION ET AJOUT AU KARDEX")
    print("-" * 40)
    
    # Choisir le mode de fonctionnement
    if os.path.exists(kardex_file):
        print("‚úÖ Fichier Kardex trouv√© - Mode AJOUT activ√©")
        success = xml_to_existing_kardex(xml_file, kardex_file)
    else:
        print("‚ö†Ô∏è  Fichier Kardex non trouv√© - Cr√©ation d'un nouveau fichier")
        success = xml_to_excel(xml_file)
    
    print("\n" + "=" * 80)
    if success:
        print("üéâ OP√âRATION TERMIN√âE AVEC SUCC√àS!")
        print(f"üìã Le fichier Kardex a √©t√© mis √† jour: {kardex_file}")
    else:
        print("‚ùå √âCHEC DE L'OP√âRATION")
    print("=" * 80)

def test_transformation_rules():
    """
    Fonction de test pour v√©rifier les r√®gles de transformation
    """
    print("üß™ Test des r√®gles de transformation:")
    print("\nüìã Nouvelles r√®gles Kardex No:")
    print("  - Location + '00'")
    print("  - Exemple: '32-30-02' ‚Üí '32-30-0200'")
    
    print("\nüìã Nouvelles r√®gles ATA:")
    print("  - Si ATAChapter existe: ATAChapter + '-00'")
    print("  - Si ATAChapter vide: extraire de Location (premier nombre avant '-')")
    print("  - Exemple: Location '32-30-02' ‚Üí ATA '32'")
    
    print("\nüìã R√®gles Item Consumed NEW:")
    print("  - TSN_Part + ' FH & ' + CSN_Part + ' FC'")
    print("  - Si une valeur manque, afficher seulement celle disponible")
    print("  - Ignorer les valeurs '0:00' et '0.00'")
    
    # Test avec des exemples
    test_cases = [
        {'Location': '72-00', 'ATAChapter': 72},
        {'Location': '32-30-01', 'ATAChapter': ''},
        {'Location': '46-10', 'ATAChapter': 46},
        {'Location': '32-30-02', 'ATAChapter': None}
    ]
    
    print("\nüî¨ Exemples de transformation:")
    for case in test_cases:
        location = case['Location']
        ata_chapter = case['ATAChapter']
        
        # Calculer Kardex No
        kardex_no = f"{location}00"
        
        # Calculer ATA
        if ata_chapter and str(ata_chapter).strip() != '':
            ata_value = f"{int(ata_chapter)}-00"
        else:
            ata_from_location = location.split('-')[0] if '-' in location else location
            ata_value = ata_from_location
        
        print(f"  Location: {location}, ATAChapter: {ata_chapter}")
        print(f"    ‚Üí Kardex No: {kardex_no}")
        print(f"    ‚Üí ATA: {ata_value}")
        print()

if __name__ == "__main__":
    # Afficher les r√®gles de transformation
    test_transformation_rules()
    print("=" * 80)
    
    # Ex√©cuter la conversion principale
    main()

# Exemples d'utilisation:
# 
# 1. Ajouter au fichier Kardex existant:
# xml_to_existing_kardex("mon_fichier.xml", "/path/to/kardex.xlsm")
#
# 2. Cr√©er un nouveau fichier:
# xml_to_excel("mon_fichier.xml", "nouveau_fichier.xlsx")
#
# 3. Ajouter √† un fichier existant quelconque:
# xml_to_excel("mon_fichier.xml", "fichier_existant.xlsx", append_to_existing=True)