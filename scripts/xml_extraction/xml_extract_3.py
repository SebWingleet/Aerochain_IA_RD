import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
import os

def xml_to_excel(xml_file_path, excel_file_path=None, append_to_existing=False):
    """
    Convertit un fichier XML avec structure RalWebDataTable en fichier Excel
    
    Args:
        xml_file_path (str): Chemin vers le fichier XML source
        excel_file_path (str): Chemin de sortie pour le fichier Excel (optionnel)
        append_to_existing (bool): Si True, ajoute aux données existantes
    
    Returns:
        str: Chemin du fichier Excel créé/modifié
    """
    
    # Si aucun chemin de sortie n'est spécifié, créer un nom basé sur le fichier XML
    if excel_file_path is None:
        base_name = os.path.splitext(xml_file_path)[0]
        excel_file_path = f"{base_name}_converted.xlsx"
    
    try:
        # Parser le fichier XML
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        
        # Liste pour stocker toutes les données
        data_list = []
        
        # Parcourir tous les éléments RalWebDataTable
        for table in root.findall('RalWebDataTable'):
            # Dictionnaire pour une ligne de données
            row_data = {}
            
            # Extraire toutes les données de chaque élément
            for child in table:
                # Nettoyer le nom de la colonne et la valeur
                column_name = child.tag.strip()
                value = child.text.strip() if child.text else ""
                row_data[column_name] = value
            
            data_list.append(row_data)
        
        # Créer un DataFrame pandas
        df = pd.DataFrame(data_list)
        
        # Nettoyer et formater les données
        df = clean_and_format_dataframe(df)
        
        # Transformer vers le format personnalisé
        transformed_df = transform_to_custom_format(df)
        
        # Gestion du fichier Excel selon le mode
        if append_to_existing and os.path.exists(excel_file_path):
            # Mode ajout : ajouter aux données existantes
            result = append_to_existing_excel(transformed_df, excel_file_path)
            if not result:
                return None
        else:
            # Mode création : créer un nouveau fichier
            create_new_excel_file(transformed_df, df, excel_file_path)
        
        print(f"✅ Conversion réussie!")
        print(f"📁 Fichier Excel: {excel_file_path}")
        if append_to_existing:
            print(f"📊 Nouvelles lignes ajoutées: {len(transformed_df)}")
        else:
            print(f"📊 Feuille 'Aircraft_Data_Custom': {len(transformed_df)} lignes, {len(transformed_df.columns)} colonnes")
            print(f"📊 Feuille 'Aircraft_Data_Original': {len(df)} lignes, {len(df.columns)} colonnes")
        
        return excel_file_path
        
    except ET.ParseError as e:
        print(f"❌ Erreur lors du parsing XML: {e}")
        return None
    except Exception as e:
        print(f"❌ Erreur générale: {e}")
        return None

def clean_and_format_dataframe(df):
    """
    Nettoie et formate le DataFrame
    
    Args:
        df (pandas.DataFrame): DataFrame à nettoyer
    
    Returns:
        pandas.DataFrame: DataFrame nettoyé
    """
    
    # Colonnes de dates à traiter
    date_columns = ['ManufactureDate', 'InstallationDate', 'FirstInstallationDate', 'ExpiryDate']
    
    for col in date_columns:
        if col in df.columns:
            # Convertir les dates au format datetime
            df[col] = pd.to_datetime(df[col], format='%d.%m.%Y', errors='coerce')
    
    # Colonnes numériques à traiter
    numeric_columns = [
        'AircraftSerialNumber', 'ModelTreeLevel', 'ATAChapter',
        'HigherAssemblyAgeingAtFitInCycles', 'ComponentAgeingatInstallationinCycles',
        'HigherAssemblyCurrentCycles', 'ComponentCurrentCycles'
    ]
    
    for col in numeric_columns:
        if col in df.columns:
            # Convertir en numérique, remplacer les erreurs par NaN
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Remplacer les chaînes vides par NaN pour une meilleure lisibilité
    df = df.replace('', pd.NA)
    
    return df

def transform_to_custom_format(df):
    """
    Transforme le DataFrame selon le format demandé avec les nouvelles colonnes
    
    Args:
        df (pandas.DataFrame): DataFrame source avec les données XML
    
    Returns:
        pandas.DataFrame: DataFrame transformé avec les nouvelles colonnes
    """
    
    # Créer un nouveau DataFrame avec les colonnes demandées
    transformed_data = []
    
    for _, row in df.iterrows():
        # Calculer les valeurs dérivées
        ata_chapter = str(row.get('ATAChapter', '')) if pd.notna(row.get('ATAChapter')) else ''
        model_tree_level = row.get('ModelTreeLevel', 0)
        
        # Calculer ATA
        ata_value = f"{ata_chapter}-00" if ata_chapter else ''
        
        # Calculer Kardex No selon les règles
        kardex_no = ''
        if ata_chapter:
            if model_tree_level == 1:
                kardex_no = f"{ata_chapter}-00-1"
            elif model_tree_level == 2:
                kardex_no = f"{ata_chapter}-00-11"
            elif model_tree_level >= 3:
                kardex_no = f"{ata_chapter}-00-111"
        
        # Calculer "Item Consumed (at installation) NEW"
        tsn_part = row.get('ComponentAgeingatInstallationinHours', '')
        csn_part = row.get('ComponentAgeingatInstallationinCycles', '')
        
        item_consumed_new = ''
        if pd.notna(tsn_part) and tsn_part != '' and pd.notna(csn_part) and csn_part != '':
            item_consumed_new = f"{tsn_part} FH & {csn_part} FC"
        elif pd.notna(tsn_part) and tsn_part != '':
            item_consumed_new = f"{tsn_part} FH"
        elif pd.notna(csn_part) and csn_part != '':
            item_consumed_new = f"{csn_part} FC"
        
        # Construire la nouvelle ligne
        new_row = {
            'Analyse': '',
            'Assembly level': '',
            'ATA': ata_value,
            'Kardex No': kardex_no,
            'Designation': row.get('InstalledPartDescription', ''),
            'P_N': row.get('InstalledManufacturerPartNumber', ''),
            'S_N': row.get('InstalledSerialNumber', ''),
            'Installation_Date_AC': row.get('FirstInstallationDate', ''),
            'TSN_AC': row.get('HigherAssemblyAgeingAtFitInHours', ''),
            'CSN_AC': row.get('HigherAssemblyAgeingAtFitInCycles', ''),
            'TSN_Part': row.get('ComponentAgeingatInstallationinHours', ''),
            'CSN_Part': row.get('ComponentAgeingatInstallationinCycles', ''),
            'Item Consumed (at installation) NEW': item_consumed_new,
            'Item Consumed (at installation) Overhaul': '',
            'Item Consumed (at installation) Maintenance': '',
            'Item Consumed (at installation) Inspection': '',
            'Monitoring': '',
            'JAA/EASA Certificate': '',
            'Item Monitoring : New (at installation) Task N': '',
            'Item Monitoring : New (at installation) Authorized': ''
        }
        
        transformed_data.append(new_row)
    
    # Créer le nouveau DataFrame
    transformed_df = pd.DataFrame(transformed_data)
    
    # Nettoyer les valeurs vides
    transformed_df = transformed_df.replace('', pd.NA)
    
    return transformed_df

def append_to_existing_excel(new_data_df, excel_file_path):
    """
    Ajoute de nouvelles données à un fichier Excel existant
    
    Args:
        new_data_df (pandas.DataFrame): Nouvelles données à ajouter
        excel_file_path (str): Chemin vers le fichier Excel existant
    
    Returns:
        bool: True si succès, False sinon
    """
    
    try:
        from openpyxl import load_workbook
        
        print(f"📖 Lecture du fichier existant: {excel_file_path}")
        
        # Charger le workbook existant
        workbook = load_workbook(excel_file_path, keep_vba=True)  # keep_vba=True pour les fichiers .xlsm
        
        # # Supposer que les données sont dans la première feuille
        # sheet_name = workbook.sheetnames[0]
        # worksheet = workbook[sheet_name]

        # Chercher la feuille "KARDEX"
        target_sheet_name = "KARDEX"
        if target_sheet_name in workbook.sheetnames:
            worksheet = workbook[target_sheet_name]
            print(f"📄 Feuille trouvée: {target_sheet_name}")
        else:
            print(f"⚠️  Feuille '{target_sheet_name}' non trouvée!")
            print(f"📋 Feuilles disponibles: {workbook.sheetnames}")
            # Utiliser la première feuille comme fallback
            sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            print(f"📄 Utilisation de la feuille par défaut: {sheet_name}")
        
        # Trouver la dernière ligne avec des données
        last_row = worksheet.max_row
        
        # Vérifier s'il y a des données existantes (en ignorant l'en-tête)
        while last_row > 1:
            # Vérifier si la ligne contient des données (au moins une cellule non vide)
            has_data = False
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=last_row, column=col).value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_data = True
                    break
            
            if has_data:
                break
            last_row -= 1
        
        print(f"📍 Dernière ligne avec données: {last_row}")
        
        # Position pour commencer à ajouter les nouvelles données
        start_row = last_row + 1
        
        print(f"📝 Ajout des nouvelles données à partir de la ligne: {start_row}")
        
        # Ajouter les nouvelles données ligne par ligne
        for idx, (_, row) in enumerate(new_data_df.iterrows(), start=start_row):
            for col_idx, value in enumerate(row.values, start=1):
                # Convertir les valeurs pandas NaN en None pour Excel
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    # Convertir les dates pandas en dates Python
                    value = value.to_pydatetime().date()
                
                worksheet.cell(row=idx, column=col_idx, value=value)
        
        # Sauvegarder le fichier
        workbook.save(excel_file_path)
        workbook.close()
        
        print(f"✅ {len(new_data_df)} nouvelles lignes ajoutées avec succès!")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors de l'ajout aux données existantes: {e}")
        return False

def create_new_excel_file(transformed_df, original_df, excel_file_path):
    """
    Crée un nouveau fichier Excel avec les données transformées
    
    Args:
        transformed_df (pandas.DataFrame): Données transformées
        original_df (pandas.DataFrame): Données originales
        excel_file_path (str): Chemin de sortie
    """
    
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        # Feuille avec les données transformées (format demandé)
        transformed_df.to_excel(writer, sheet_name='Aircraft_Data_Custom', index=False)
        
        # Feuille avec les données originales (pour référence)
        original_df.to_excel(writer, sheet_name='Aircraft_Data_Original', index=False)
        
        # Obtenir les feuilles de calcul pour le formatage
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Définir une largeur maximale raisonnable
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

def xml_to_existing_kardex(xml_file_path, kardex_file_path):
    """
    Fonction spécialisée pour ajouter des données XML au fichier Kardex existant
    
    Args:
        xml_file_path (str): Chemin vers le fichier XML
        kardex_file_path (str): Chemin vers le fichier Kardex (.xlsm)
    
    Returns:
        bool: True si succès, False sinon
    """
    
    print(f"🎯 Mode Kardex: Ajout au fichier existant")
    print(f"📁 XML source: {xml_file_path}")
    print(f"📁 Kardex cible: {kardex_file_path}")
    
    # Vérifier que les fichiers existent
    if not os.path.exists(xml_file_path):
        print(f"❌ Fichier XML non trouvé: {xml_file_path}")
        return False
    
    if not os.path.exists(kardex_file_path):
        print(f"❌ Fichier Kardex non trouvé: {kardex_file_path}")
        return False
    
    # Utiliser la fonction principale avec l'option append
    result = xml_to_excel(xml_file_path, kardex_file_path, append_to_existing=True)
    
    return result is not None

def analyze_xml_structure(xml_file_path):
    """
    Analyse la structure du fichier XML et affiche des informations utiles
    
    Args:
        xml_file_path (str): Chemin vers le fichier XML
    """
    
    try:
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        
        print("🔍 Analyse de la structure XML:")
        print(f"Élément racine: {root.tag}")
        
        # Compter les tables
        tables = root.findall('RalWebDataTable')
        print(f"Nombre de RalWebDataTable: {len(tables)}")
        
        # Analyser la première table pour voir la structure
        if tables:
            first_table = tables[0]
            print(f"\n📋 Colonnes disponibles ({len(first_table)} au total):")
            
            for i, child in enumerate(first_table, 1):
                value = child.text if child.text else "(vide)"
                print(f"  {i:2d}. {child.tag:35} = {value}")
        
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {e}")

def main():
    """
    Fonction principale pour exécution en tant que script
    """
    
    # Configuration des chemins
    xml_file = "/Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/INPUT_DOCS/xml/H160-MIS-DATA-PACK-1054-Applied Configuration.xml"  # Remplacez par le chemin de votre fichier XML
    kardex_file = "/Users/sebastienbatty/Documents/1_Wingleet/2_DEV/THC/Output_docs/10_KARDEX_H160_MSN1054_V3.xlsm"
    
    print("=" * 80)
    print("🚁 CONVERTISSEUR XML VERS KARDEX H160")
    print("=" * 80)
    
    # Vérifier si le fichier XML existe
    if not os.path.exists(xml_file):
        print(f"❌ Fichier XML non trouvé: {xml_file}")
        print("💡 Modifiez la variable 'xml_file' avec le bon chemin")
        return
    
    # Analyser la structure (optionnel)
    print("🔍 ANALYSE DE LA STRUCTURE XML")
    print("-" * 40)
    analyze_xml_structure(xml_file)
    
    print("\n🔄 CONVERSION ET AJOUT AU KARDEX")
    print("-" * 40)
    
    # Choisir le mode de fonctionnement
    if os.path.exists(kardex_file):
        print("✅ Fichier Kardex trouvé - Mode AJOUT activé")
        success = xml_to_existing_kardex(xml_file, kardex_file)
    else:
        print("⚠️  Fichier Kardex non trouvé - Création d'un nouveau fichier")
        success = xml_to_excel(xml_file)
    
    print("\n" + "=" * 80)
    if success:
        print("🎉 OPÉRATION TERMINÉE AVEC SUCCÈS!")
        print(f"📋 Le fichier Kardex a été mis à jour: {kardex_file}")
    else:
        print("❌ ÉCHEC DE L'OPÉRATION")
    print("=" * 80)

def test_transformation_rules():
    """
    Fonction de test pour vérifier les règles de transformation
    """
    print("🧪 Test des règles de transformation:")
    print("\n📋 Règles Kardex No:")
    print("  - ModelTreeLevel == 1  → ATAChapter-00-1")
    print("  - ModelTreeLevel == 2  → ATAChapter-00-11") 
    print("  - ModelTreeLevel >= 3  → ATAChapter-00-111")
    
    print("\n📋 Règles Item Consumed NEW:")
    print("  - TSN_Part + ' FH & ' + CSN_Part + ' FC'")
    print("  - Si une valeur manque, afficher seulement celle disponible")
    
    # Test avec des exemples
    test_cases = [
        {'ATAChapter': 72, 'ModelTreeLevel': 1},
        {'ATAChapter': 72, 'ModelTreeLevel': 2},
        {'ATAChapter': 32, 'ModelTreeLevel': 4},
        {'ATAChapter': 46, 'ModelTreeLevel': 5}
    ]
    
    print("\n🔬 Exemples de transformation Kardex No:")
    for case in test_cases:
        ata = case['ATAChapter']
        level = case['ModelTreeLevel']
        if level == 1:
            result = f"{ata}-00-1"
        elif level == 2:
            result = f"{ata}-00-11"
        else:
            result = f"{ata}-00-111"
        print(f"  ATA: {ata}, Level: {level} → {result}")

if __name__ == "__main__":
    # Afficher les règles de transformation
    test_transformation_rules()
    print("=" * 80)
    
    # Exécuter la conversion principale
    main()

# Exemples d'utilisation:
# 
# 1. Ajouter au fichier Kardex existant:
# xml_to_existing_kardex("mon_fichier.xml", "/path/to/kardex.xlsm")
#
# 2. Créer un nouveau fichier:
# xml_to_excel("mon_fichier.xml", "nouveau_fichier.xlsx")
#
# 3. Ajouter à un fichier existant quelconque:
# xml_to_excel("mon_fichier.xml", "fichier_existant.xlsx", append_to_existing=True)